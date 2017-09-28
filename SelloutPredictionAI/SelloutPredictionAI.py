
from numpy import exp, array, random, dot
from openpyxl import load_workbook

class NeuralNetwork():
    def __init__(self):
        # Seed the random number generator, so it generates the same numbers
        # every time the program runs.
        random.seed(1)

        # We model a single neuron, with 3 input connections and 1 output connection.
        # We assign random weights to a 3 x 1 matrix, with values in the range -1 to 1
        # and mean 0.
        self.synaptic_weights = 2 * random.random((3, 1)) - 1

    # The Sigmoid function, which describes an S shaped curve.
    # We pass the weighted sum of the inputs through this function to
    # normalise them between 0 and 1.
    def __sigmoid(self, x):
        return 1 / (1 + exp(-x))

    # The derivative of the Sigmoid function.
    # This is the gradient of the Sigmoid curve.
    # It indicates how confident we are about the existing weight.
    def __sigmoid_derivative(self, x):
        return x * (1 - x)

    # We train the neural network through a process of trial and error.
    # Adjusting the synaptic weights each time.
    def train(self, training_set_inputs, training_set_outputs, number_of_training_iterations):
        for iteration in range(number_of_training_iterations):
            # Pass the training set through our neural network (a single neuron).
            output = self.think(training_set_inputs)

            # Calculate the error (The difference between the desired output
            # and the predicted output).
            error = training_set_outputs - output

            # Multiply the error by the input and again by the gradient of the Sigmoid curve.
            # This means less confident weights are adjusted more.
            # This means inputs, which are zero, do not cause changes to the weights.
            adjustment = dot(training_set_inputs.T, error * self.__sigmoid_derivative(output))

            # Adjust the weights.
            self.synaptic_weights += adjustment

    # The neural network thinks.
    def think(self, inputs):
        # Pass inputs through our neural network (our single neuron).
        return self.__sigmoid(dot(inputs, self.synaptic_weights))

def Normalize(val, min, max):
    normalized = (val - min)/(max - min)
    return normalized

if __name__ == "__main__":

    
    neural_network = NeuralNetwork()


    training_set_inputs = []
    training_set_outputs = []

    wb = load_workbook("Data.xlsx")
    ws = wb.active

    cols = []
    for c in range(1,4):
        data = []
        for r in range(2,13):
            data.append(ws.cell(row = r, column = c).value)
        data1 = map(float,data)
        normalized = []
        max_num = max(data1)
        min_num = min(data1)
        for d in data1:
            normalized.append(Normalize(d,min_num,max_num))
        cols.append(normalized)


    for i in range(len(cols[0])):
        show = []
        for c in cols:
            show.append(c[i])
        training_set_inputs.append(show)
    training_set_inputs_array = array(training_set_inputs)

    results = []
    for r in range(2,13):
        results.append(ws.cell(row = r, column = 4).value)

    training_set_outputs = map(float, results)
    training_set_outputs_array = array([training_set_outputs]).T


    print "Random Weights: \n"
    print neural_network.synaptic_weights
    print "\n"

    neural_network.train(training_set_inputs_array, training_set_outputs_array, 50)

    print("New synaptic weights after training: ")
    print neural_network.synaptic_weights
    print
    "\n"

    print("Considering new situation [1, 1, 0] -> ?: ")
    print(neural_network.think(array([1, 1, 0])))
